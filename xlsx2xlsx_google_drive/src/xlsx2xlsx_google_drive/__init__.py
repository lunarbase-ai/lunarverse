import os
import io
import re
import pandas as pd
from typing import Optional

from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaFileUpload
from google.oauth2.service_account import Credentials
from openpyxl import load_workbook
import json

from lunarcore.component.lunar_component import LunarComponent
from lunarcore.component.component_group import ComponentGroup
from lunarcore.component.data_types import DataType


class Xlsx2XlsxGoogleDrive(
    LunarComponent,
    component_name="xlsx2xlsx Google Drive",
    component_description="Updates XLSX files in Google Drive by downloading, updating, and re-uploading files from Google Drive.",
    input_types={"file_link": DataType.TEXT, "credentials_json": DataType.TEXT, "excel_file": DataType.TEXT, "sheet_name": DataType.TEXT},
    output_type=DataType.TEXT,
    component_group=ComponentGroup.UTILS,
):

    def __init__(self, **kwargs):
        super().__init__(configuration=kwargs)

    def extract_file_id(self, file_link: str) -> str:
        match = re.search(r'/d/([a-zA-Z0-9_-]+)', file_link)
        if not match:
            raise ValueError("Invalid Google Drive link")
        return match.group(1)

    def run(self, file_link: str, credentials_json: str, excel_file: str, sheet_name: str) -> str:
        SCOPES = ["https://www.googleapis.com/auth/drive"]
        creds = Credentials.from_service_account_file(str(credentials_json), scopes=SCOPES)

        service = build("drive", "v3", credentials=creds)

        file_id = self.extract_file_id(file_link)
        file_metadata = service.files().get(fileId=file_id).execute()
        file_name = file_metadata['name']

        request = service.files().get_media(fileId=file_id)
        file = io.BytesIO()
        downloader = MediaIoBaseDownload(file, request)

        done = False
        while not done:
            status, done = downloader.next_chunk()

        file.seek(0)
        output_file_path = f"/tmp/{file_name}"
        with open(output_file_path, "wb") as f:
            f.write(file.read())

        df = pd.read_excel(excel_file, usecols=lambda column: not column.startswith('Unnamed'))
        dicionario = df.to_dict()

        dicionario_celulas = {}

        colunas = list(dicionario.keys())

        for indice_col, nome_coluna in enumerate(colunas):
            letra_coluna = chr(ord('A') + indice_col)
            dicionario_celulas[f"{letra_coluna}1"] = nome_coluna
            
            for indice_linha, valor in dicionario[nome_coluna].items():
                dicionario_celulas[f"{letra_coluna}{indice_linha + 2}"] = valor

        wb = load_workbook(output_file_path)
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
        else:
            ws = wb[sheet_name]

        for cell, value in dicionario_celulas.items():
            ws[cell].value = value

        temp_updated_path = "/tmp/temp_updated.xlsx"
        wb.save(temp_updated_path)

        media = MediaFileUpload(temp_updated_path,
                                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                resumable=True)

        updated_file = service.files().update(
            fileId=file_id,
            media_body=media,
            fields="id, name"
        ).execute()        

        return updated_file['name']